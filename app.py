from flask import (
    Flask, request, redirect, url_for, render_template_string,
    session, jsonify, flash, send_file
)
from functools import wraps
from uuid import uuid4
from datetime import datetime
from werkzeug.utils import secure_filename
from sqlalchemy import create_engine, text
from sqlalchemy.exc import SQLAlchemyError
import boto3
import os
import io
import re

from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.environ.get("APP_SECRET_KEY", "doi-secret-key-rat-dai-o-day")

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("Thiếu DATABASE_URL trong Environment Variables.")

engine = create_engine(
    DATABASE_URL,
    pool_pre_ping=True,
    future=True,
)

EXPORT_FILE = "du_lieu_ma_vach.xlsx"
LOCAL_UPLOAD_FOLDER = "uploads_barcode_file"
os.makedirs(LOCAL_UPLOAD_FOLDER, exist_ok=True)

ALLOWED_BARCODE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".pdf"}

DEFAULT_ROWS = [
    {
        "id": str(uuid4()),
        "ten_thung": "Viên treo bồn cầu",
        "so_luong_san_pham_thung": "9 vỉ/thùng",
        "ten": "EverShine - Viên treo bồn cầu hương Chanh",
        "ma_vach_sp": "8935367400014",
        "so_ma_vach_thung": "38935367400015",
        "updated_by": "Super Admin",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    },
    {
        "id": str(uuid4()),
        "ten_thung": "Viên treo bồn cầu",
        "so_luong_san_pham_thung": "9 vỉ/thùng",
        "ten": "EverShine - Viên treo bồn cầu hương Lavender",
        "ma_vach_sp": "8935367400021",
        "so_ma_vach_thung": "38935367400022",
        "updated_by": "Super Admin",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    },
]

DEFAULT_SETTINGS = {
    "group_pattern": "3-5-5-1",
}


def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def format_grouped_number(value: str, pattern: str) -> str:
    digits = only_digits(value)
    if not digits:
        return value or ""
    try:
        groups = [int(x) for x in str(pattern).split("-") if x.strip()]
    except ValueError:
        groups = []
    if not groups:
        return digits

    out = []
    idx = 0
    for size in groups:
        if idx >= len(digits):
            break
        out.append(digits[idx:idx + size])
        idx += size
    if idx < len(digits):
        out.append(digits[idx:])
    return " ".join([x for x in out if x])


def allowed_barcode_file(filename: str) -> bool:
    ext = os.path.splitext(filename.lower())[1]
    return ext in ALLOWED_BARCODE_EXTENSIONS


def barcode_file_type(filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]
    return "pdf" if ext == ".pdf" else "image"


def db_execute(query: str, params=None, fetch=False, fetchone=False):
    params = params or {}
    with engine.begin() as conn:
        result = conn.execute(text(query), params)
        if fetchone:
            row = result.mappings().first()
            return dict(row) if row else None
        if fetch:
            return [dict(r) for r in result.mappings().all()]
        return None


def r2_enabled() -> bool:
    return all([
        os.environ.get("R2_ACCOUNT_ID"),
        os.environ.get("R2_ACCESS_KEY_ID"),
        os.environ.get("R2_SECRET_ACCESS_KEY"),
        os.environ.get("R2_BUCKET"),
    ])


def get_s3():
    return boto3.client(
        service_name="s3",
        endpoint_url=f"https://{os.environ['R2_ACCOUNT_ID']}.r2.cloudflarestorage.com",
        aws_access_key_id=os.environ["R2_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["R2_SECRET_ACCESS_KEY"],
        region_name="auto",
    )


def storage_mode() -> str:
    return "r2" if r2_enabled() else "local"


def init_db():
    db_execute("""
        CREATE TABLE IF NOT EXISTS admins (
            id TEXT PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT,
            role TEXT,
            display_name TEXT,
            created_at TEXT
        )
    """)

    db_execute("""
        CREATE TABLE IF NOT EXISTS barcode_rows (
            id TEXT PRIMARY KEY,
            ten_thung TEXT,
            so_luong_san_pham_thung TEXT,
            ten TEXT,
            ma_vach_sp TEXT,
            so_ma_vach_thung TEXT,
            updated_by TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    db_execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)

    db_execute("""
        CREATE TABLE IF NOT EXISTS row_barcode_files (
            id TEXT PRIMARY KEY,
            row_id TEXT UNIQUE,
            original_name TEXT,
            stored_key TEXT,
            file_type TEXT,
            storage_mode TEXT,
            uploaded_by TEXT,
            created_at TEXT
        )
    """)

    admin_count = db_execute("SELECT COUNT(*) AS total FROM admins", fetchone=True)
    if admin_count and admin_count["total"] == 0:
        now = now_str()
        for row in [
            (str(uuid4()), "superadmin", "123456", "super_admin", "Super Admin", now),
            (str(uuid4()), "admin1", "123456", "admin", "Admin 1", now),
            (str(uuid4()), "admin2", "123456", "admin", "Admin 2", now),
        ]:
            db_execute("""
                INSERT INTO admins (id, username, password, role, display_name, created_at)
                VALUES (:id, :username, :password, :role, :display_name, :created_at)
            """, {
                "id": row[0],
                "username": row[1],
                "password": row[2],
                "role": row[3],
                "display_name": row[4],
                "created_at": row[5],
            })

    db_execute("""
        UPDATE admins
        SET role = 'super_admin'
        WHERE username = 'superadmin' AND (role IS NULL OR role = '')
    """)
    db_execute("""
        UPDATE admins
        SET role = 'admin'
        WHERE role IS NULL OR role = ''
    """)
    db_execute("""
        UPDATE admins
        SET display_name = username
        WHERE display_name IS NULL OR display_name = ''
    """)

    for key, value in DEFAULT_SETTINGS.items():
        db_execute("""
            INSERT INTO app_settings (key, value)
            VALUES (:key, :value)
            ON CONFLICT (key) DO NOTHING
        """, {"key": key, "value": value})

    row_count = db_execute("SELECT COUNT(*) AS total FROM barcode_rows", fetchone=True)
    if row_count and row_count["total"] == 0:
        for row in DEFAULT_ROWS:
            db_execute("""
                INSERT INTO barcode_rows (
                    id, ten_thung, so_luong_san_pham_thung, ten,
                    ma_vach_sp, so_ma_vach_thung, updated_by, created_at, updated_at
                ) VALUES (
                    :id, :ten_thung, :so_luong_san_pham_thung, :ten,
                    :ma_vach_sp, :so_ma_vach_thung, :updated_by, :created_at, :updated_at
                )
            """, row)


def get_settings():
    rows = db_execute("SELECT key, value FROM app_settings", fetch=True) or []
    settings = {r["key"]: r["value"] for r in rows}
    merged = DEFAULT_SETTINGS.copy()
    merged.update(settings)
    return merged


def list_barcode_file_map():
    rows = db_execute("SELECT * FROM row_barcode_files", fetch=True) or []
    return {r["row_id"]: r for r in rows}


def list_rows(search=""):
    if search.strip():
        like = f"%{search.strip()}%"
        rows = db_execute("""
            SELECT * FROM barcode_rows
            WHERE ten_thung LIKE :like
               OR so_luong_san_pham_thung LIKE :like
               OR ten LIKE :like
               OR ma_vach_sp LIKE :like
               OR so_ma_vach_thung LIKE :like
               OR updated_by LIKE :like
            ORDER BY ten_thung ASC, ten ASC
        """, {"like": like}, fetch=True) or []
    else:
        rows = db_execute("""
            SELECT * FROM barcode_rows
            ORDER BY ten_thung ASC, ten ASC
        """, fetch=True) or []

    settings = get_settings()
    barcode_map = list_barcode_file_map()

    for row in rows:
        row["ma_vach_sp_grouped"] = format_grouped_number(row["ma_vach_sp"], settings["group_pattern"])
        row["so_ma_vach_thung_grouped"] = format_grouped_number(row["so_ma_vach_thung"], settings["group_pattern"])
        row["barcode_file"] = barcode_map.get(row["id"])
        row["has_barcode_file"] = row["id"] in barcode_map
        row["barcode_file_type"] = barcode_map.get(row["id"], {}).get("file_type", "")
    return rows


def get_row(row_id):
    return db_execute(
        "SELECT * FROM barcode_rows WHERE id = :row_id",
        {"row_id": row_id},
        fetchone=True,
    )


def create_row(form, updated_by):
    now = now_str()
    db_execute("""
        INSERT INTO barcode_rows (
            id, ten_thung, so_luong_san_pham_thung, ten,
            ma_vach_sp, so_ma_vach_thung, updated_by, created_at, updated_at
        ) VALUES (
            :id, :ten_thung, :so_luong_san_pham_thung, :ten,
            :ma_vach_sp, :so_ma_vach_thung, :updated_by, :created_at, :updated_at
        )
    """, {
        "id": str(uuid4()),
        "ten_thung": form.get("ten_thung", "").strip(),
        "so_luong_san_pham_thung": form.get("so_luong_san_pham_thung", "").strip(),
        "ten": form.get("ten", "").strip(),
        "ma_vach_sp": only_digits(form.get("ma_vach_sp", "")),
        "so_ma_vach_thung": only_digits(form.get("so_ma_vach_thung", "")),
        "updated_by": updated_by,
        "created_at": now,
        "updated_at": now,
    })


def update_row_db(row_id, form, updated_by):
    db_execute("""
        UPDATE barcode_rows
        SET ten_thung = :ten_thung,
            so_luong_san_pham_thung = :so_luong_san_pham_thung,
            ten = :ten,
            ma_vach_sp = :ma_vach_sp,
            so_ma_vach_thung = :so_ma_vach_thung,
            updated_by = :updated_by,
            updated_at = :updated_at
        WHERE id = :row_id
    """, {
        "ten_thung": form.get("ten_thung", "").strip(),
        "so_luong_san_pham_thung": form.get("so_luong_san_pham_thung", "").strip(),
        "ten": form.get("ten", "").strip(),
        "ma_vach_sp": only_digits(form.get("ma_vach_sp", "")),
        "so_ma_vach_thung": only_digits(form.get("so_ma_vach_thung", "")),
        "updated_by": updated_by,
        "updated_at": now_str(),
        "row_id": row_id,
    })


def get_row_barcode_file(row_id):
    return db_execute(
        "SELECT * FROM row_barcode_files WHERE row_id = :row_id",
        {"row_id": row_id},
        fetchone=True,
    )


def delete_storage_object(stored_key: str, mode: str):
    if not stored_key:
        return
    if mode == "r2" and r2_enabled():
        try:
            s3 = get_s3()
            s3.delete_object(Bucket=os.environ["R2_BUCKET"], Key=stored_key)
        except Exception:
            pass
    else:
        local_path = os.path.join(LOCAL_UPLOAD_FOLDER, stored_key)
        if os.path.exists(local_path):
            try:
                os.remove(local_path)
            except OSError:
                pass


def save_row_barcode_file(row_id, file_storage, uploaded_by):
    old = get_row_barcode_file(row_id)
    if old:
        delete_storage_object(old["stored_key"], old.get("storage_mode", "local"))

    original_name = file_storage.filename or "barcode_file"
    safe_original = secure_filename(original_name)
    file_id = str(uuid4())
    ext = os.path.splitext(safe_original)[1].lower()
    file_type = barcode_file_type(original_name)
    mode = storage_mode()

    if mode == "r2":
        stored_key = f"barcode/{row_id}/{file_id}{ext}"
        s3 = get_s3()
        file_storage.stream.seek(0)
        s3.upload_fileobj(
            file_storage.stream,
            os.environ["R2_BUCKET"],
            stored_key,
            ExtraArgs={"ContentType": file_storage.mimetype or "application/octet-stream"},
        )
    else:
        stored_key = f"{file_id}_{safe_original}"
        file_storage.save(os.path.join(LOCAL_UPLOAD_FOLDER, stored_key))

    db_execute("""
        INSERT INTO row_barcode_files (
            id, row_id, original_name, stored_key, file_type, storage_mode, uploaded_by, created_at
        ) VALUES (
            :id, :row_id, :original_name, :stored_key, :file_type, :storage_mode, :uploaded_by, :created_at
        )
        ON CONFLICT (row_id) DO UPDATE SET
            original_name = EXCLUDED.original_name,
            stored_key = EXCLUDED.stored_key,
            file_type = EXCLUDED.file_type,
            storage_mode = EXCLUDED.storage_mode,
            uploaded_by = EXCLUDED.uploaded_by,
            created_at = EXCLUDED.created_at
    """, {
        "id": file_id,
        "row_id": row_id,
        "original_name": original_name,
        "stored_key": stored_key,
        "file_type": file_type,
        "storage_mode": mode,
        "uploaded_by": uploaded_by,
        "created_at": now_str(),
    })


def delete_row_barcode_file(row_id):
    row = get_row_barcode_file(row_id)
    if not row:
        return
    delete_storage_object(row["stored_key"], row.get("storage_mode", "local"))
    db_execute("DELETE FROM row_barcode_files WHERE row_id = :row_id", {"row_id": row_id})


def delete_row_db(row_id):
    delete_row_barcode_file(row_id)
    db_execute("DELETE FROM barcode_rows WHERE id = :row_id", {"row_id": row_id})


def replace_all_rows(rows, updated_by):
    # Xóa file barcode cũ
    old_files = db_execute("SELECT * FROM row_barcode_files", fetch=True) or []
    for f in old_files:
        delete_storage_object(f["stored_key"], f.get("storage_mode", "local"))

    db_execute("DELETE FROM row_barcode_files")
    db_execute("DELETE FROM barcode_rows")

    for row in rows:
        now = now_str()
        db_execute("""
            INSERT INTO barcode_rows (
                id, ten_thung, so_luong_san_pham_thung, ten,
                ma_vach_sp, so_ma_vach_thung, updated_by, created_at, updated_at
            ) VALUES (
                :id, :ten_thung, :so_luong_san_pham_thung, :ten,
                :ma_vach_sp, :so_ma_vach_thung, :updated_by, :created_at, :updated_at
            )
        """, {
            "id": str(uuid4()),
            "ten_thung": row.get("ten_thung", "").strip(),
            "so_luong_san_pham_thung": row.get("so_luong_san_pham_thung", "").strip(),
            "ten": row.get("ten", "").strip(),
            "ma_vach_sp": only_digits(row.get("ma_vach_sp", "")),
            "so_ma_vach_thung": only_digits(row.get("so_ma_vach_thung", "")),
            "updated_by": updated_by,
            "created_at": now,
            "updated_at": now,
        })


def group_rows(rows):
    grouped = []
    current = None
    for row in rows:
        key = (row["ten_thung"], row["so_luong_san_pham_thung"])
        if current and current["key"] == key:
            current["items"].append(row)
        else:
            current = {
                "key": key,
                "ten_thung": row["ten_thung"],
                "so_luong_san_pham_thung": row["so_luong_san_pham_thung"],
                "items": [row],
            }
            grouped.append(current)
    return grouped


def list_admins():
    return db_execute("""
        SELECT id, username, password, role, display_name, created_at
        FROM admins
        ORDER BY CASE WHEN role='super_admin' THEN 0 ELSE 1 END, username ASC
    """, fetch=True) or []


def admin_exists(username):
    row = db_execute(
        "SELECT id FROM admins WHERE lower(username) = lower(:username)",
        {"username": username.strip()},
        fetchone=True,
    )
    return row is not None


def verify_admin(username, password):
    return db_execute("""
        SELECT id, username, role, display_name
        FROM admins
        WHERE username = :username AND password = :password
    """, {
        "username": username.strip(),
        "password": password,
    }, fetchone=True)


def get_admin_by_id(admin_id):
    return db_execute("""
        SELECT id, username, role, display_name, created_at
        FROM admins
        WHERE id = :admin_id
    """, {"admin_id": admin_id}, fetchone=True)


def create_admin_user(username, password, role, display_name):
    db_execute("""
        INSERT INTO admins (id, username, password, role, display_name, created_at)
        VALUES (:id, :username, :password, :role, :display_name, :created_at)
    """, {
        "id": str(uuid4()),
        "username": username.strip(),
        "password": password,
        "role": role,
        "display_name": display_name.strip(),
        "created_at": now_str(),
    })


def delete_admin_user(admin_id):
    db_execute("DELETE FROM admins WHERE id = :admin_id", {"admin_id": admin_id})


def update_admin_password(admin_id, new_password):
    db_execute("""
        UPDATE admins
        SET password = :password
        WHERE id = :admin_id
    """, {
        "password": new_password,
        "admin_id": admin_id,
    })


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Bạn cần đăng nhập admin để chỉnh sửa.", "error")
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


PAGE_HTML = """
<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Bảng mã vạch nội bộ</title>
  <script src="https://unpkg.com/html5-qrcode" type="text/javascript"></script>
  <style>
    :root { --bg:#eef2f7; --panel:#fff; --line:#cfd8e3; --line-dark:#94a3b8; --text:#0f172a; --muted:#475569; --brand:#0f766e; --brand-dark:#115e59; --brand-soft:#ecfdf5; --danger:#c62828; --warning:#b45309; --info:#175cd3; --shadow:0 12px 30px rgba(15,23,42,.08); --radius:18px; }
    *{box-sizing:border-box}
    body{margin:0;background:linear-gradient(180deg,#edf3f9 0%,#f8fafc 100%);color:var(--text);font-family:system-ui,-apple-system,"Segoe UI",Roboto,Arial,Helvetica,sans-serif}
    .container{max-width:1680px;margin:0 auto;padding:20px}
    .topbar{display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap;margin-bottom:18px}
    h1{margin:0;font-size:30px}
    .sub{margin-top:6px;color:var(--muted);font-size:14px}
    .card{background:var(--panel);border:1px solid var(--line);border-radius:var(--radius);box-shadow:var(--shadow);padding:18px;margin-bottom:16px}
    .toolbar,.toolbar2{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap}
    .toolbar2{margin-top:12px}
    .btn{border:0;background:var(--brand);color:#fff;padding:10px 16px;border-radius:12px;cursor:pointer;text-decoration:none;font-size:14px;font-weight:700;white-space:nowrap;box-shadow:0 6px 16px rgba(15,118,110,.18)}
    .btn:hover{background:var(--brand-dark)}
    .btn.secondary{background:#fff;color:var(--text);border:1px solid var(--line);box-shadow:none}
    .btn.danger{background:var(--danger);box-shadow:none}
    .btn.warning{background:var(--warning);box-shadow:none}
    .btn.download{background:#1d4ed8;box-shadow:none}
    .badge{display:inline-flex;align-items:center;gap:6px;background:var(--brand-soft);color:#047857;font-size:12px;font-weight:800;padding:7px 12px;border-radius:999px;border:1px solid #bbf7d0}
    .badge.viewer{background:#eff6ff;color:var(--info);border-color:#bfdbfe}
    .flash{border-radius:12px;padding:12px 14px;margin-bottom:10px;font-size:14px;background:#edfdf2;color:#067647;border:1px solid #ccebcf}
    .flash.error{background:#fff1f1;color:#b42318;border-color:#f0c7c7}
    .grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px}
    .grid-1{grid-column:1 / -1}
    label{display:block;margin-bottom:6px;font-size:14px;font-weight:800;color:#1e293b}
    input[type='text'],input[type='password'],input[type='file'],select{width:100%;border:1px solid var(--line);border-radius:12px;padding:11px 12px;font-size:14px;background:#fff;font-family:inherit}
    .table-wrap{overflow-x:auto;border:1px solid var(--line-dark);background:#fff;border-radius:14px}
    table{width:100%;border-collapse:collapse;min-width:1600px;table-layout:fixed;background:#fff}
    th,td{border:1px solid var(--line-dark);padding:12px 10px;text-align:center;vertical-align:middle;font-size:14px;line-height:1.35}
    th{background:#e2e8f0;font-weight:800;color:#0f172a}
    td.left{text-align:left}
    td.group-cell,td.qty-cell{font-size:18px;line-height:1.45;font-weight:700;background:#fafafa}
    .barcode-action-box{display:flex;flex-direction:column;gap:8px;align-items:center;justify-content:center}
    .status-ok{font-weight:800;color:#067647}
    .status-no{font-weight:800;color:#b42318}
    .status-pdf{font-weight:800;color:#1d4ed8}
    .actions{display:flex;gap:8px;flex-wrap:wrap;justify-content:center}
    .login-box{max-width:450px;margin:40px auto}
    .scan-grid{display:grid;grid-template-columns:360px 1fr;gap:18px;align-items:start}
    #reader{width:100%;max-width:360px;min-height:250px;border:1px solid var(--line);border-radius:16px;overflow:hidden;background:#fff}
    .scan-result{border:2px dashed #94a3b8;border-radius:16px;padding:16px 18px;background:linear-gradient(180deg,#fff 0%,#f8fafc 100%);font-size:18px;font-weight:700;color:#0f172a}
    .scan-label{color:var(--muted);font-size:14px;font-weight:600;margin-top:6px}
    .hit-card{border:2px solid #dbe3ec;border-radius:16px;padding:18px 20px;margin-top:12px;background:linear-gradient(180deg,#fff 0%,#f8fafc 100%);font-size:18px;line-height:1.55;font-weight:600;box-shadow:0 8px 20px rgba(15,23,42,.04)}
    .hit-title{font-size:30px;line-height:1.2;font-weight:900;color:#0f172a;margin-bottom:10px}
    .hit-row{margin-top:4px;font-size:20px}
    .hit-code{font-size:28px;font-weight:900;color:#0f172a;letter-spacing:.02em}
    .hit-code.red{color:#b91c1c}
    .hit-meta{color:var(--muted);font-size:15px;font-weight:600;margin-top:6px}
    .meta{color:var(--muted);font-size:13px}
    .stats-pill{display:inline-flex;align-items:center;gap:8px;padding:8px 12px;border-radius:999px;background:#f8fafc;border:1px solid var(--line);font-size:13px;font-weight:700;color:#334155}
    .section-title{font-size:18px;font-weight:900;color:#0f172a;margin:0}
    @media print{.topbar,.toolbar form,.btn,.scan-grid,.actions,.barcode-action-box form{display:none !important}body{background:#fff}.card{box-shadow:none;border:1px solid #d1d5db}}
    @media (max-width:1100px){.grid,.scan-grid{grid-template-columns:1fr}td.group-cell,td.qty-cell{font-size:16px}.hit-title{font-size:24px}.hit-row{font-size:18px}.hit-code{font-size:24px}}
  </style>
</head>
<body>
  <div class="container">
    <div class="topbar">
      <div><h1>Bảng mã vạch nội bộ</h1><div class="sub">PostgreSQL + upload barcode bền hơn, hỗ trợ ảnh hoặc PDF.</div></div>
      <div style="display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
        {% if is_admin %}
          <span class="badge">{{ admin_display_name }}{% if is_super_admin %} - Super Admin{% else %} - Admin{% endif %}</span>
          <a class="btn secondary" href="{{ url_for('logout') }}">Đăng xuất</a>
        {% else %}
          <span class="badge viewer">Chế độ xem</span>
          <a class="btn" href="{{ url_for('login') }}">Đăng nhập admin</a>
        {% endif %}
      </div>
    </div>

    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="flash {% if category == 'error' %}error{% endif %}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}

    {% if show_login %}
      <div class="card login-box">
        <h2 class="section-title">Đăng nhập admin</h2>
        <form method="post" action="{{ url_for('login') }}" style="margin-top:14px;">
          <label>Tài khoản</label><input type="text" name="username" required>
          <label style="margin-top:10px;">Mật khẩu</label><input type="password" name="password" required>
          <div style="margin-top:14px; display:flex; gap:8px;">
            <button class="btn" type="submit">Đăng nhập</button>
            <a class="btn secondary" href="{{ url_for('index') }}">Quay lại</a>
          </div>
        </form>
      </div>
    {% else %}

      <div class="card">
        <div class="toolbar">
          <div>
            <h2 class="section-title">Tìm nhanh / scan mã</h2>
            <div class="meta">Quét hoặc nhập mã để xem kết quả lớn, rõ.</div>
          </div>
          <form method="get" action="{{ url_for('index') }}" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center;">
            <input type="text" name="q" value="{{ q }}" placeholder="Tìm theo tên, mã..." style="width:360px; max-width:100%;">
            <button class="btn secondary" type="submit">Tìm</button>
            {% if q %}<a class="btn secondary" href="{{ url_for('index') }}">Xoá lọc</a>{% endif %}
          </form>
        </div>
        <div class="scan-grid" style="margin-top:16px;">
          <div>
            <div id="reader"></div>
            <div class="toolbar2">
              <button class="btn" type="button" onclick="startScanner()">Bật camera scan</button>
              <button class="btn secondary" type="button" onclick="stopScanner()">Tắt camera</button>
              <button class="btn secondary" type="button" onclick="switchCamera()">Đổi camera</button>
              <button class="btn secondary" type="button" onclick="toggleTorch()">Đèn flash</button>
            </div>
            <div class="meta" id="cameraStatus" style="margin-top:8px;">Chưa bật camera</div>
          </div>
          <div>
            <div class="scan-result">
              <div><strong>Mã vừa quét:</strong> <span id="scanText">Chưa có</span></div>
              <div class="scan-label">Không có camera thì nhập tay hoặc dùng máy quét USB.</div>
            </div>
            <div id="scanHits"></div>
          </div>
        </div>
      </div>

      {% if is_admin %}
      <div class="card">
        <div class="toolbar">
          <div>
            <h2 class="section-title">Quản lý admin</h2>
            <div class="meta">Super Admin được tạo, xoá admin và đổi mọi mật khẩu. Admin thường chỉ được đổi mật khẩu của chính mình.</div>
          </div>
          <div class="stats-pill">Tổng tài khoản quản trị: {{ admins|length }}</div>
        </div>
        <div class="grid" style="margin-top:14px;">
          <div>
            {% if is_super_admin %}
              <form method="post" action="{{ url_for('add_admin') }}">
                <label>Tên hiển thị</label><input type="text" name="display_name" required>
                <label style="margin-top:10px;">Tài khoản admin mới</label><input type="text" name="username" required>
                <label style="margin-top:10px;">Mật khẩu</label><input type="text" name="password" required>
                <label style="margin-top:10px;">Quyền</label>
                <select name="role">
                  <option value="admin">Admin thường</option>
                  <option value="super_admin">Super Admin</option>
                </select>
                <div style="margin-top:12px;"><button class="btn" type="submit">Tạo admin</button></div>
              </form>
            {% else %}
              <div class="hit-card" style="margin-top:0;">
                <div class="hit-title" style="font-size:22px;">Bạn là admin thường</div>
                <div class="hit-meta">Bạn chỉ được đổi mật khẩu của chính mình.</div>
              </div>
            {% endif %}
          </div>
          <div>
            <div class="meta" style="margin-bottom:8px;">Danh sách admin</div>
            {% for admin in admins %}
              <div class="hit-card" style="margin-top:0; margin-bottom:10px; font-size:16px;">
                <div style="font-size:22px; font-weight:900;">{{ admin['display_name'] }}</div>
                <div class="hit-meta">Tài khoản: {{ admin['username'] }} | Quyền: {{ 'Super Admin' if admin['role'] == 'super_admin' else 'Admin' }}</div>
                <div class="hit-meta">Tạo lúc: {{ admin['created_at'] }}</div>
                {% if is_super_admin or session_admin_id == admin['id'] %}
                  <form method="post" action="{{ url_for('change_admin_password', admin_id=admin['id']) }}" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:12px;">
                    <input type="text" name="new_password" required placeholder="Mật khẩu mới" style="max-width:240px;">
                    <button class="btn secondary" type="submit">Đổi mật khẩu</button>
                  </form>
                {% endif %}
                {% if is_super_admin and admins|length > 1 %}
                  <form method="post" action="{{ url_for('delete_admin', admin_id=admin['id']) }}" onsubmit="return confirm('Xoá admin này?');" style="margin-top:10px;">
                    <button class="btn danger" type="submit">Xoá admin</button>
                  </form>
                {% endif %}
              </div>
            {% endfor %}
          </div>
        </div>
      </div>

      <div class="card">
        <div class="toolbar">
          <div>
            <h2 class="section-title">Import / Export dữ liệu</h2>
            <div class="meta">Import và export bằng file Excel .xlsx.</div>
          </div>
          <a class="btn secondary" href="{{ url_for('export_xlsx') }}">Export XLSX</a>
        </div>
        <form method="post" action="{{ url_for('import_xlsx') }}" enctype="multipart/form-data" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:12px;">
          <input type="file" name="file" accept=".xlsx" required style="max-width:380px;">
          <button class="btn warning" type="submit">Import XLSX</button>
        </form>
      </div>

      <div class="card">
        <div class="toolbar">
          <div>
            <h2 class="section-title">{% if edit_row %}Sửa dòng{% else %}Thêm dòng mới{% endif %}</h2>
            <div class="meta">Mã thùng nhập tay theo đúng mã bạn muốn.</div>
          </div>
          {% if edit_row %}<a class="btn secondary" href="{{ url_for('index') }}">Huỷ sửa</a>{% endif %}
        </div>
        <form method="post" action="{% if edit_row %}{{ url_for('update_row', row_id=edit_row['id']) }}{% else %}{{ url_for('add_row') }}{% endif %}" style="margin-top:12px;">
          <div class="grid">
            <div><label>Tên thùng</label><input type="text" name="ten_thung" value="{{ edit_row['ten_thung'] if edit_row else '' }}" required></div>
            <div><label>Số lượng sản phẩm/thùng</label><input type="text" name="so_luong_san_pham_thung" value="{{ edit_row['so_luong_san_pham_thung'] if edit_row else '' }}" required></div>
            <div class="grid-1"><label>Tên</label><input type="text" name="ten" value="{{ edit_row['ten'] if edit_row else '' }}" required></div>
            <div><label>Mã vạch SP</label><input type="text" name="ma_vach_sp" value="{{ edit_row['ma_vach_sp'] if edit_row else '' }}" required></div>
            <div><label>Số mã vạch thùng</label><input type="text" name="so_ma_vach_thung" value="{{ edit_row['so_ma_vach_thung'] if edit_row else '' }}" required></div>
          </div>
          <div style="margin-top:14px; display:flex; gap:8px; flex-wrap:wrap;">
            <button class="btn" type="submit">{% if edit_row %}Lưu thay đổi{% else %}Thêm dòng{% endif %}</button>
          </div>
        </form>
      </div>
      {% endif %}

      <div class="card">
        <div class="toolbar">
          <div>
            <h2 class="section-title">Bảng dữ liệu</h2>
            <div class="meta">Tổng dòng hiển thị: {{ rows|length }}</div>
          </div>
          <div style="display:flex; gap:8px; flex-wrap:wrap;">
            <div class="stats-pill">Người sửa hiển thị trên từng dòng</div>
            <button class="btn secondary" type="button" onclick="window.print()">In bảng</button>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead>
              <tr>
                <th style="width:11%;">Tên Thùng</th>
                <th style="width:9%;">Số lượng SP/thùng</th>
                <th style="width:17%;">Tên</th>
                <th style="width:9%;">Mã vạch SP</th>
                <th style="width:11%;">Số mã vạch thùng</th>
                <th style="width:21%;">Mã vạch</th>
                <th style="width:10%;">Người sửa</th>
                {% if is_admin %}<th style="width:12%;">Thao tác</th>{% endif %}
              </tr>
            </thead>
            <tbody>
              {% for group in grouped_rows %}
                {% for row in group['items'] %}
                  <tr>
                    {% if loop.index0 == 0 %}
                      <td class="group-cell" rowspan="{{ group['items']|length }}">{{ group['ten_thung'] }}</td>
                      <td class="qty-cell" rowspan="{{ group['items']|length }}">{{ group['so_luong_san_pham_thung'] }}</td>
                    {% endif %}
                    <td class="left">{{ row['ten'] }}</td>
                    <td>{{ row['ma_vach_sp_grouped'] }}</td>
                    <td>{{ row['so_ma_vach_thung_grouped'] }}</td>
                    <td>
                      <div class="barcode-action-box">
                        {% if row['has_barcode_file'] %}
                          {% if row['barcode_file_type'] == 'pdf' %}
                            <div class="status-pdf">Đã có barcode PDF</div>
                          {% else %}
                            <div class="status-ok">Đã có barcode ảnh</div>
                          {% endif %}
                          <a class="btn download" href="{{ url_for('download_row_barcode', row_id=row['id']) }}">Tải xuống</a>
                        {% else %}
                          <div class="status-no">Chưa có barcode</div>
                        {% endif %}

                        {% if is_admin %}
                          <form method="post" action="{{ url_for('upload_row_barcode', row_id=row['id']) }}" enctype="multipart/form-data" style="display:flex; gap:6px; flex-direction:column; width:100%;">
                            <input type="file" name="barcode_file" accept=".png,.jpg,.jpeg,.webp,.pdf" required>
                            <button class="btn" type="submit">Tải lên</button>
                          </form>

                          {% if row['has_barcode_file'] %}
                          <form method="post" action="{{ url_for('delete_row_barcode', row_id=row['id']) }}" onsubmit="return confirm('Xoá barcode của dòng này?');">
                            <button class="btn danger" type="submit">Xoá barcode</button>
                          </form>
                          {% endif %}
                        {% endif %}
                      </div>
                    </td>
                    <td>{{ row['updated_by'] or '' }}</td>
                    {% if is_admin %}
                      <td>
                        <div class="actions">
                          <a class="btn secondary" href="{{ url_for('edit_row', row_id=row['id']) }}">Sửa</a>
                          <form method="post" action="{{ url_for('delete_row', row_id=row['id']) }}" onsubmit="return confirm('Xoá dòng này?');">
                            <button class="btn danger" type="submit">Xoá</button>
                          </form>
                        </div>
                      </td>
                    {% endif %}
                  </tr>
                {% endfor %}
              {% else %}
                <tr><td colspan="8">Không có dữ liệu phù hợp.</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    {% endif %}
  </div>

  <script>
    const allRows = {{ rows_json | safe }};
    const SETTINGS = {{ settings_json | safe }};
    let scannerInstance = null;
    let scanning = false;
    let currentCameraIndex = 0;
    let cameraList = [];
    let lastScannedText = '';
    let lastScanAt = 0;
    let torchEnabled = false;
    let beepCtx = null;

    function digitsOnly(v){
      return String(v || '').replace(/\\D/g, '');
    }

    function groupNumber(value, pattern){
      const digits = digitsOnly(value);
      const groups = String(pattern || '').split('-').map(x => parseInt(x, 10)).filter(x => !Number.isNaN(x) && x > 0);
      if (!groups.length) return digits;
      let out = [];
      let idx = 0;
      groups.forEach(size => {
        if (idx < digits.length) {
          out.push(digits.slice(idx, idx + size));
          idx += size;
        }
      });
      if (idx < digits.length) out.push(digits.slice(idx));
      return out.join(' ');
    }

    function findMatches(code) {
      const normalized = (code || '').trim().toLowerCase();
      if (!normalized) return [];
      return allRows.filter(row =>
        String(row.ma_vach_sp || '').includes(normalized) ||
        String(row.so_ma_vach_thung || '').includes(normalized) ||
        String(row.ten || '').toLowerCase().includes(normalized) ||
        String(row.ten_thung || '').toLowerCase().includes(normalized)
      );
    }

    function renderHits(code) {
      const wrap = document.getElementById('scanHits');
      document.getElementById('scanText').innerText = code || 'Chưa có';
      const hits = findMatches(code);
      if (!code) {
        wrap.innerHTML = '';
        return;
      }
      if (!hits.length) {
        wrap.innerHTML = '<div class="hit-card"><div class="hit-title" style="font-size:24px;">Không tìm thấy dữ liệu phù hợp</div><div class="hit-meta">Hãy kiểm tra lại mã quét hoặc nhập tay vào ô tìm kiếm.</div></div>';
        return;
      }
      wrap.innerHTML = hits.map(row => `
        <div class="hit-card">
          <div class="hit-title">${row.ten}</div>
          <div class="hit-row">Thùng: <strong>${row.ten_thung}</strong> | SL/thùng: <strong>${row.so_luong_san_pham_thung}</strong></div>
          <div class="hit-row">Mã SP: <span class="hit-code">${groupNumber(row.ma_vach_sp, SETTINGS.group_pattern)}</span></div>
          <div class="hit-row">Mã thùng: <span class="hit-code red">${groupNumber(row.so_ma_vach_thung, SETTINGS.group_pattern)}</span></div>
          <div class="hit-row">Người sửa: <strong>${row.updated_by || ''}</strong></div>
          <div class="hit-row">Trạng thái barcode: <strong>${row.has_barcode_file ? (row.barcode_file_type === 'pdf' ? 'Đã có barcode PDF' : 'Đã có barcode ảnh') : 'Chưa có barcode'}</strong></div>
        </div>`).join('');
    }

    function setCameraStatus(text) {
      const el = document.getElementById('cameraStatus');
      if (el) el.innerText = text;
    }

    function beepOk() {
      try {
        if (!beepCtx) beepCtx = new (window.AudioContext || window.webkitAudioContext)();
        const osc = beepCtx.createOscillator();
        const gain = beepCtx.createGain();
        osc.type = 'sine';
        osc.frequency.value = 880;
        gain.gain.value = 0.05;
        osc.connect(gain);
        gain.connect(beepCtx.destination);
        osc.start();
        setTimeout(() => osc.stop(), 120);
      } catch (e) {}
    }

    function onScanSuccess(decodedText) {
      const now = Date.now();
      if (decodedText === lastScannedText && now - lastScanAt < 1500) return;
      lastScannedText = decodedText;
      lastScanAt = now;
      renderHits(decodedText);
      const searchInput = document.querySelector('input[name="q"]');
      if (searchInput) searchInput.value = decodedText;
      beepOk();
      setCameraStatus('Đã quét: ' + decodedText);
    }

    async function startScanner() {
      if (scanning) return;
      try {
        cameraList = await Html5Qrcode.getCameras();
        if (!cameraList || !cameraList.length) {
          alert('Không tìm thấy camera trên thiết bị này.');
          setCameraStatus('Không tìm thấy camera');
          return;
        }
        if (currentCameraIndex >= cameraList.length) currentCameraIndex = 0;
        const cameraId = cameraList[currentCameraIndex].id;
        scannerInstance = new Html5Qrcode('reader');
        await scannerInstance.start(
          cameraId,
          {
            fps: 12,
            qrbox: function(viewfinderWidth, viewfinderHeight) {
              const w = Math.floor(Math.min(viewfinderWidth * 0.88, 420));
              const h = Math.floor(Math.min(viewfinderHeight * 0.34, 180));
              return { width: w, height: h };
            },
            aspectRatio: 1.777,
            disableFlip: false,
            experimentalFeatures: { useBarCodeDetectorIfSupported: true },
            formatsToSupport: [
              Html5QrcodeSupportedFormats.CODE_128,
              Html5QrcodeSupportedFormats.CODE_39,
              Html5QrcodeSupportedFormats.EAN_13,
              Html5QrcodeSupportedFormats.EAN_8,
              Html5QrcodeSupportedFormats.UPC_A,
              Html5QrcodeSupportedFormats.UPC_E,
              Html5QrcodeSupportedFormats.QR_CODE
            ]
          },
          onScanSuccess,
          () => {}
        );
        scanning = true;
        torchEnabled = false;
        const camName = cameraList[currentCameraIndex].label || `Camera ${currentCameraIndex + 1}`;
        setCameraStatus('Đang dùng: ' + camName);
      } catch (e) {
        alert('Không bật được camera. Hãy kiểm tra quyền camera hoặc thử Chrome/Edge trên điện thoại.');
        setCameraStatus('Bật camera thất bại');
      }
    }

    async function stopScanner() {
      try {
        if (!scannerInstance || !scanning) return;
        await scannerInstance.stop();
        await scannerInstance.clear();
      } catch (e) {}
      scannerInstance = null;
      scanning = false;
      torchEnabled = false;
      setCameraStatus('Đã tắt camera');
    }

    async function switchCamera() {
      try {
        if (!cameraList || !cameraList.length) cameraList = await Html5Qrcode.getCameras();
        if (!cameraList.length) {
          alert('Không có camera để đổi.');
          return;
        }
        currentCameraIndex = (currentCameraIndex + 1) % cameraList.length;
        if (scanning) {
          await stopScanner();
          await startScanner();
        } else {
          const camName = cameraList[currentCameraIndex].label || `Camera ${currentCameraIndex + 1}`;
          setCameraStatus('Đã chọn: ' + camName);
        }
      } catch (e) {
        alert('Không đổi được camera.');
      }
    }

    async function toggleTorch() {
      if (!scannerInstance || !scanning) {
        alert('Hãy bật camera trước.');
        return;
      }
      try {
        const capabilities = scannerInstance.getRunningTrackCapabilities ? scannerInstance.getRunningTrackCapabilities() : null;
        if (!capabilities || !capabilities.torch) {
          alert('Camera này không hỗ trợ đèn flash.');
          return;
        }
        torchEnabled = !torchEnabled;
        await scannerInstance.applyVideoConstraints({ advanced: [{ torch: torchEnabled }] });
        setCameraStatus(torchEnabled ? 'Đã bật đèn flash' : 'Đã tắt đèn flash');
      } catch (e) {
        alert('Không bật được đèn flash trên thiết bị này.');
      }
    }

    {% if q %}renderHits({{ q | tojson }});{% endif %}
  </script>
</body>
</html>
"""


init_db()


def render_page(rows, q, is_admin, show_login, edit_row):
    settings = get_settings()
    return render_template_string(
        PAGE_HTML,
        rows=rows,
        grouped_rows=group_rows(rows),
        rows_json=rows,
        q=q,
        settings=settings,
        settings_json=settings,
        is_admin=is_admin,
        is_super_admin=session.get("admin_role") == "super_admin",
        admin_display_name=session.get("admin_display_name", ""),
        session_admin_id=session.get("admin_id", ""),
        show_login=show_login,
        edit_row=edit_row,
        admins=list_admins(),
    )


@app.route("/")
def index():
    q = request.args.get("q", "")
    rows = list_rows(q)
    return render_page(rows, q, session.get("is_admin", False), False, None)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        admin = verify_admin(username, password)
        if admin:
            session["is_admin"] = True
            session["admin_id"] = admin["id"]
            session["admin_username"] = admin["username"]
            session["admin_role"] = admin["role"]
            session["admin_display_name"] = admin["display_name"]
            flash("Đăng nhập admin thành công.", "success")
            return redirect(request.args.get("next") or url_for("index"))
        flash("Sai tài khoản hoặc mật khẩu admin.", "error")
    return render_page(list_rows(), "", False, True, None)


@app.route("/logout")
def logout():
    session.clear()
    flash("Đã đăng xuất.", "success")
    return redirect(url_for("index"))


@app.route("/add", methods=["POST"])
@admin_required
def add_row():
    if not request.form.get("ten", "").strip():
        flash("Tên sản phẩm không được để trống.", "error")
        return redirect(url_for("index"))
    create_row(request.form, session.get("admin_display_name", session.get("admin_username", "")))
    flash("Đã thêm dòng mới.", "success")
    return redirect(url_for("index"))


@app.route("/edit/<row_id>")
@admin_required
def edit_row(row_id):
    row = get_row(row_id)
    if not row:
        flash("Không tìm thấy dòng cần sửa.", "error")
        return redirect(url_for("index"))
    return render_page(list_rows(), "", True, False, row)


@app.route("/update/<row_id>", methods=["POST"])
@admin_required
def update_row(row_id):
    if not get_row(row_id):
        flash("Dòng này không còn tồn tại.", "error")
        return redirect(url_for("index"))
    update_row_db(row_id, request.form, session.get("admin_display_name", session.get("admin_username", "")))
    flash("Đã cập nhật dòng.", "success")
    return redirect(url_for("index"))


@app.route("/delete/<row_id>", methods=["POST"])
@admin_required
def delete_row(row_id):
    delete_row_db(row_id)
    flash("Đã xoá dòng.", "success")
    return redirect(url_for("index"))


@app.route("/export")
@admin_required
def export_xlsx():
    rows = list_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Du lieu ma vach"
    headers = ["ten_thung", "so_luong_san_pham_thung", "ten", "ma_vach_sp", "so_ma_vach_thung", "updated_by"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["ten_thung"],
            row["so_luong_san_pham_thung"],
            row["ten"],
            row["ma_vach_sp"],
            row["so_ma_vach_thung"],
            row.get("updated_by", ""),
        ])

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 42)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=EXPORT_FILE,
    )


@app.route("/import", methods=["POST"])
@admin_required
def import_xlsx():
    file = request.files.get("file")
    if not file or not file.filename.lower().endswith(".xlsx"):
        flash("Hãy chọn file XLSX hợp lệ.", "error")
        return redirect(url_for("index"))

    try:
        wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash("Không đọc được file XLSX.", "error")
        return redirect(url_for("index"))

    rows_iter = list(ws.iter_rows(values_only=True))
    if not rows_iter:
        flash("File XLSX không có dữ liệu.", "error")
        return redirect(url_for("index"))

    headers = ["" if h is None else str(h).strip() for h in rows_iter[0]]
    required_cols = ["ten_thung", "so_luong_san_pham_thung", "ten", "ma_vach_sp", "so_ma_vach_thung"]
    if not all(col in headers for col in required_cols):
        flash("File XLSX thiếu cột bắt buộc.", "error")
        return redirect(url_for("index"))

    index_map = {name: headers.index(name) for name in required_cols}
    imported_rows = []
    for excel_row in rows_iter[1:]:
        ten_value = excel_row[index_map["ten"]] if index_map["ten"] < len(excel_row) else ""
        if ten_value is None or str(ten_value).strip() == "":
            continue
        imported_rows.append({
            "ten_thung": "" if excel_row[index_map["ten_thung"]] is None else str(excel_row[index_map["ten_thung"]]).strip(),
            "so_luong_san_pham_thung": "" if excel_row[index_map["so_luong_san_pham_thung"]] is None else str(excel_row[index_map["so_luong_san_pham_thung"]]).strip(),
            "ten": str(ten_value).strip(),
            "ma_vach_sp": "" if excel_row[index_map["ma_vach_sp"]] is None else str(excel_row[index_map["ma_vach_sp"]]).strip(),
            "so_ma_vach_thung": "" if excel_row[index_map["so_ma_vach_thung"]] is None else str(excel_row[index_map["so_ma_vach_thung"]]).strip(),
        })

    replace_all_rows(imported_rows, session.get("admin_display_name", session.get("admin_username", "")))
    flash(f"Đã import {len(imported_rows)} dòng từ XLSX.", "success")
    return redirect(url_for("index"))


@app.route("/row-barcode/<row_id>/upload", methods=["POST"])
@admin_required
def upload_row_barcode(row_id):
    if not get_row(row_id):
        flash("Không tìm thấy dòng sản phẩm.", "error")
        return redirect(url_for("index"))

    file = request.files.get("barcode_file")
    if not file or not file.filename:
        flash("Hãy chọn file barcode.", "error")
        return redirect(url_for("index"))

    if not allowed_barcode_file(file.filename):
        flash("Chỉ cho phép file PNG/JPG/JPEG/WEBP/PDF.", "error")
        return redirect(url_for("index"))

    try:
        save_row_barcode_file(row_id, file, session.get("admin_display_name", session.get("admin_username", "")))
        if storage_mode() == "r2":
            flash("Đã tải lên barcode lên cloud.", "success")
        else:
            flash("Đã tải lên barcode local. Chưa cấu hình cloud nên file có thể không bền sau redeploy.", "success")
    except Exception as e:
        flash(f"Lỗi upload barcode: {e}", "error")
    return redirect(url_for("index"))


@app.route("/row-barcode/<row_id>/download")
def download_row_barcode(row_id):
    row = get_row_barcode_file(row_id)
    if not row:
        flash("Sản phẩm này chưa có barcode.", "error")
        return redirect(url_for("index"))

    if row.get("storage_mode") == "r2" and r2_enabled():
        try:
            s3 = get_s3()
            url = s3.generate_presigned_url(
                "get_object",
                Params={"Bucket": os.environ["R2_BUCKET"], "Key": row["stored_key"]},
                ExpiresIn=3600,
            )
            return redirect(url)
        except Exception:
            flash("Không tạo được link tải file cloud.", "error")
            return redirect(url_for("index"))

    file_path = os.path.join(LOCAL_UPLOAD_FOLDER, row["stored_key"])
    if not os.path.exists(file_path):
        flash("File barcode không còn tồn tại.", "error")
        return redirect(url_for("index"))

    return send_file(file_path, as_attachment=True, download_name=row["original_name"])


@app.route("/row-barcode/<row_id>/delete", methods=["POST"])
@admin_required
def delete_row_barcode(row_id):
    delete_row_barcode_file(row_id)
    flash("Đã xoá barcode của sản phẩm.", "success")
    return redirect(url_for("index"))


@app.route("/admins/add", methods=["POST"])
@admin_required
def add_admin():
    if session.get("admin_role") != "super_admin":
        flash("Chỉ Super Admin mới được tạo admin.", "error")
        return redirect(url_for("index"))
    display_name = request.form.get("display_name", "").strip()
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    role = request.form.get("role", "admin").strip()
    if not display_name or not username or not password:
        flash("Tên hiển thị, tài khoản và mật khẩu không được để trống.", "error")
        return redirect(url_for("index"))
    if admin_exists(username):
        flash("Tài khoản admin đã tồn tại.", "error")
        return redirect(url_for("index"))
    if role not in ["admin", "super_admin"]:
        role = "admin"
    create_admin_user(username, password, role, display_name)
    flash("Đã tạo admin mới.", "success")
    return redirect(url_for("index"))


@app.route("/admins/<admin_id>/password", methods=["POST"])
@admin_required
def change_admin_password(admin_id):
    new_password = request.form.get("new_password", "").strip()
    if session.get("admin_role") != "super_admin" and session.get("admin_id") != admin_id:
        flash("Admin thường chỉ được đổi mật khẩu của chính mình.", "error")
        return redirect(url_for("index"))
    if not new_password:
        flash("Mật khẩu mới không được để trống.", "error")
        return redirect(url_for("index"))
    update_admin_password(admin_id, new_password)
    flash("Đã đổi mật khẩu admin.", "success")
    return redirect(url_for("index"))


@app.route("/admins/<admin_id>/delete", methods=["POST"])
@admin_required
def delete_admin(admin_id):
    if session.get("admin_role") != "super_admin":
        flash("Chỉ Super Admin mới được xoá admin.", "error")
        return redirect(url_for("index"))
    admins = list_admins()
    if len(admins) <= 1:
        flash("Phải giữ ít nhất 1 admin.", "error")
        return redirect(url_for("index"))
    target = get_admin_by_id(admin_id)
    if target and target.get("role") == "super_admin":
        super_admin_count = sum(1 for a in admins if a.get("role") == "super_admin")
        if super_admin_count <= 1:
            flash("Phải giữ ít nhất 1 Super Admin.", "error")
            return redirect(url_for("index"))
    delete_admin_user(admin_id)
    if session.get("admin_id") == admin_id:
        session.clear()
        flash("Bạn đã tự xoá tài khoản đang đăng nhập.", "success")
        return redirect(url_for("index"))
    flash("Đã xoá admin.", "success")
    return redirect(url_for("index"))


@app.route("/api/rows")
def api_rows():
    return jsonify(list_rows())


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)